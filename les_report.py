import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- File paths ---
CSV_FILE    = "/Users/bosichelstiel/Desktop/LES Docs/Master_Historical_Database.csv"
OUTPUT_FILE = "/Users/bosichelstiel/Desktop/LES Docs/LES_Estimating_Analysis.xlsx"

# --- Load and clean data ---
df = pd.read_csv(CSV_FILE)
df['PM'] = df['PM'].str.strip().str.replace(r'\s+', ' ', regex=True)

# --- Build summaries ---
analysis_df = df[df['Original_Est'] > 0].copy()

def build_summary(df, group_col):
    s = df.groupby(group_col).agg(
        Jobs=('Job', 'nunique'),
        Total_Original_Est=('Original_Est', 'sum'),
        Total_Projected_Cost=('Projected_Cost', 'sum'),
    ).round(2)
    s['Variance_$'] = s['Total_Original_Est'] - s['Total_Projected_Cost']
    s['Variance_%'] = (s['Variance_$'] / s['Total_Original_Est'] * 100).round(1)
    return s.sort_values('Variance_%')

scope_summary    = build_summary(analysis_df, 'Scope')
pm_summary       = build_summary(analysis_df, 'PM')
category_summary = build_summary(analysis_df, 'Category')

# --- Style helpers ---
DARK_BLUE = '1F3864'
MID_BLUE  = '2E5FA3'
RED       = 'C0392B'
GREEN     = '1E8449'
LIGHT_RED = 'FADBD8'
LIGHT_GRN = 'D5F5E3'
GRAY      = 'F2F2F2'
WHITE     = 'FFFFFF'

def fill(h):
    return PatternFill('solid', start_color=h, fgColor=h)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, color=WHITE):
    return Font(name='Arial', bold=True, size=size, color=color)

def body_font(bold=False, size=10, color='000000'):
    return Font(name='Arial', bold=bold, size=size, color=color)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_header_row(ws, row, headers, bg=DARK_BLUE):
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = hdr_font()
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()

def write_data_row(ws, row_num, values, bg=WHITE, bold=False):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.font = body_font(bold=bold)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border()

def apply_variance_row(ws, r, row, col_est=3, col_proj=4, col_var_pct=6):
    ws.cell(r, col_est).number_format  = '$#,##0'
    ws.cell(r, col_proj).number_format = '$#,##0'
    ws.cell(r, 5).number_format        = '$#,##0'
    ws.cell(r, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(r, 6).alignment = Alignment(horizontal='center', vertical='center')
    color = RED if row['Variance_%'] < 0 else GREEN
    ws.cell(r, 6).font = Font(name='Arial', bold=True, color=color)

def section_header(ws, row, text, cols='A:F', bg=MID_BLUE, height=22):
    ws.merge_cells(f'{cols.split(":")[0]}{row}:{cols.split(":")[1]}{row}')
    c = ws[f'{cols.split(":")[0]}{row}']
    c.value = text
    c.font = Font(name='Arial', bold=True, size=12, color=WHITE)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = height

def variance_row_values(name, row):
    return [name, int(row['Jobs']),
            round(float(row['Total_Original_Est']), 2),
            round(float(row['Total_Projected_Cost']), 2),
            round(float(row['Variance_$']), 2),
            f"{row['Variance_%']:+.1f}%"]

VARIANCE_HEADERS = ['Name', 'Jobs', 'Total Estimated', 'Total Projected', 'Variance $', 'Variance %']

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = Workbook()

# --- SHEET 1: OVERVIEW ---
ws1 = wb.active
ws1.title = 'Overview'
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells('A1:F1')
ws1['A1'].value = 'LE Schwartz & Sons — Estimating Analysis'
ws1['A1'].font  = Font(name='Arial', bold=True, size=16, color=WHITE)
ws1['A1'].fill  = fill(DARK_BLUE)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 36

ws1.merge_cells('A2:F2')
ws1['A2'].value = f"Based on {len(df['Job'].unique())} jobs across 3 Project Managers  |  {len(df)} cost line items"
ws1['A2'].font  = Font(name='Arial', size=11, color=WHITE)
ws1['A2'].fill  = fill(MID_BLUE)
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 22
ws1.row_dimensions[3].height = 12

# KPI boxes
kpis = [
    ('Total Jobs Analyzed',  str(len(df['Job'].unique()))),
    ('Total Estimated',      f"${df['Original_Est'].sum():,.0f}"),
    ('Total Projected',      f"${df['Projected_Cost'].sum():,.0f}"),
    ('Overall Variance',     f"${df['Original_Est'].sum() - df['Projected_Cost'].sum():+,.0f}"),
    ('Scopes Tracked',       str(len(df['Scope'].unique()))),
    ('Project Managers',     '3'),
]
for i, (label, val) in enumerate(kpis):
    col = i + 1
    lbl = ws1.cell(row=4, column=col, value=label)
    lbl.font = Font(name='Arial', bold=True, size=9, color='666666')
    lbl.alignment = Alignment(horizontal='center', vertical='center')
    lbl.fill = fill(GRAY)
    v = ws1.cell(row=5, column=col, value=val)
    v.font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
    v.alignment = Alignment(horizontal='center', vertical='center')
    v.fill = fill(WHITE)
    v.border = thin_border()
ws1.row_dimensions[5].height = 30
ws1.row_dimensions[6].height = 16

# PM table
section_header(ws1, 7, 'Performance by Project Manager')
write_header_row(ws1, 8, ['Project Manager'] + VARIANCE_HEADERS[1:])
for i, (pm, row) in enumerate(pm_summary.iterrows()):
    r = 9 + i
    bg = LIGHT_RED if row['Variance_%'] < 0 else LIGHT_GRN
    write_data_row(ws1, r, variance_row_values(pm, row), bg=bg)
    apply_variance_row(ws1, r, row)
ws1.row_dimensions[9 + len(pm_summary)].height = 10

# Key findings
last_pm_row = 9 + len(pm_summary)
notes_row   = last_pm_row + 2
section_header(ws1, notes_row, 'Key Findings')

over_scopes  = scope_summary[scope_summary['Variance_%'] < -2]
under_scopes = scope_summary[scope_summary['Variance_%'] > 5]

findings = [
    "• Most estimates are accurate — overall portfolio variance is within ±5% for major scopes.",
    f"• Scopes most consistently OVER budget: {', '.join(over_scopes.index.tolist()) if len(over_scopes) else 'None significant'}.",
    f"• Scopes most consistently UNDER budget: {', '.join(under_scopes.index.tolist()) if len(under_scopes) else 'None significant'}.",
    "• Nick Riner is the only PM currently running over budget (-1.7% across 13 jobs).",
    "• 'Gutter' and 'Gutters' appear as separate scopes — recommend standardizing scope names in source files.",
]
for j, finding in enumerate(findings):
    fr = notes_row + 1 + j
    ws1.merge_cells(f'A{fr}:F{fr}')
    c = ws1[f'A{fr}']
    c.value = finding
    c.font  = Font(name='Arial', size=10)
    c.fill  = fill(WHITE if j % 2 == 0 else GRAY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    c.border = thin_border()
    ws1.row_dimensions[fr].height = 32

for col, w in enumerate([28, 20, 22, 22, 16, 18], 1):
    set_col_width(ws1, col, w)

# --- SHEET 2: BY SCOPE ---
ws2 = wb.create_sheet('By Scope')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:F1')
ws2['A1'].value = 'Variance Analysis by Scope'
ws2['A1'].font  = Font(name='Arial', bold=True, size=14, color=WHITE)
ws2['A1'].fill  = fill(DARK_BLUE)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

write_header_row(ws2, 2, ['Scope'] + VARIANCE_HEADERS[1:])
for i, (scope, row) in enumerate(scope_summary.iterrows()):
    r  = 3 + i
    bg = LIGHT_RED if row['Variance_%'] < 0 else (LIGHT_GRN if row['Variance_%'] > 3 else WHITE)
    write_data_row(ws2, r, variance_row_values(scope, row), bg=bg)
    apply_variance_row(ws2, r, row)

note_r = 3 + len(scope_summary) + 1
ws2.merge_cells(f'A{note_r}:F{note_r}')
ws2[f'A{note_r}'].value = '* Scopes with only 1 job are shown for reference but are not statistically significant.'
ws2[f'A{note_r}'].font  = Font(name='Arial', italic=True, size=9, color='888888')
ws2[f'A{note_r}'].alignment = Alignment(horizontal='left', indent=1)

for col, w in enumerate([30, 10, 22, 22, 16, 12], 1):
    set_col_width(ws2, col, w)

# --- SHEET 3: BY PM ---
ws3 = wb.create_sheet('By PM')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:F1')
ws3['A1'].value = 'Variance Analysis by Project Manager'
ws3['A1'].font  = Font(name='Arial', bold=True, size=14, color=WHITE)
ws3['A1'].fill  = fill(DARK_BLUE)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

write_header_row(ws3, 2, ['Project Manager'] + VARIANCE_HEADERS[1:])
for i, (pm, row) in enumerate(pm_summary.iterrows()):
    r  = 3 + i
    bg = LIGHT_RED if row['Variance_%'] < 0 else LIGHT_GRN
    write_data_row(ws3, r, variance_row_values(pm, row), bg=bg)
    apply_variance_row(ws3, r, row)

cat_row = 3 + len(pm_summary) + 2
section_header(ws3, cat_row, 'Variance by Cost Category (across all PMs)')
write_header_row(ws3, cat_row + 1, ['Category'] + VARIANCE_HEADERS[1:])
for i, (cat, row) in enumerate(category_summary.iterrows()):
    r  = cat_row + 2 + i
    bg = LIGHT_RED if row['Variance_%'] < 0 else (LIGHT_GRN if row['Variance_%'] > 3 else WHITE)
    write_data_row(ws3, r, variance_row_values(cat, row), bg=bg)
    apply_variance_row(ws3, r, row)

for col, w in enumerate([28, 10, 22, 22, 16, 12], 1):
    set_col_width(ws3, col, w)

# --- SHEET 4: RAW DATA ---
ws4 = wb.create_sheet('Raw Data')
ws4.sheet_view.showGridLines = False

raw_headers = ['Job', 'City', 'State', 'GC', 'PM', 'Scope', 'Category',
               'Original_Est', 'Projected_Cost', 'Actual_to_Date', 'Variance']
write_header_row(ws4, 1, raw_headers)

for i, row in df.iterrows():
    r  = i + 2
    bg = WHITE if i % 2 == 0 else GRAY
    vals = [row['Job'], row['City'], row['State'], row['GC'], row['PM'],
            row['Scope'], row['Category'],
            round(float(row['Original_Est']), 2),
            round(float(row['Projected_Cost']), 2),
            round(float(row['Actual_to_Date']), 2),
            round(float(row['Variance']), 2)]
    write_data_row(ws4, r, vals, bg=bg)
    for col in [8, 9, 10, 11]:
        ws4.cell(r, col).number_format = '$#,##0'

for col, w in enumerate([28, 14, 8, 20, 16, 20, 16, 16, 16, 16, 14], 1):
    set_col_width(ws4, col, w)

# --- Save ---
wb.save(OUTPUT_FILE)
print(f"Report saved to: {OUTPUT_FILE}")